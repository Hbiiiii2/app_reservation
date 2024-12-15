from datetime import datetime
import pandas as pd
import os
import pytz
import random
import string
from openpyxl import Workbook, load_workbook

# WARNA
def warna(teks, warna):
    warna_dict = {
        'merah': "\033[91m",
        'hijau': "\033[92m",
        'kuning': "\033[93m",
        'biru': "\033[94m",
        'cyan': "\033[96m",
        'putih': "\033[97m",
        'reset': "\033[0m"
    }
    
    return f"{warna_dict.get(warna, warna_dict['reset'])}{teks}{warna_dict['reset']}"

def generate_userid():
    letters = ''.join(random.choices(string.ascii_uppercase, k=3))
    numbers = ''.join(random.choices(string.digits, k=4))
    return letters + numbers

def load_members():
    if os.path.exists(file_excel):
        wb = load_workbook(file_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Mulai dari baris kedua untuk menghindari header
            nama, userid, no_telepon, discount_rate = row
            akun[nama] = {
                'userid': userid,
                'discount_rate': discount_rate,
                'no_telepon': no_telepon
            }
    else:
        print(warna("File data_member.xlsx tidak ditemukan. Tidak ada data member yang dimuat.", 'merah'))

def hitung_total(harga, nama_member=None):
    pajak = harga * 0.05
    service = harga * 0.02
    diskon = 0
    
    if nama_member and nama_member in akun:
        discount_rate = akun[nama_member]['discount_rate'] / 100
        diskon = harga * discount_rate
        
    total = harga + pajak + service - diskon
    return {
        'subtotal': harga,
        'pajak': pajak,
        'service': service,
        'diskon': diskon,
        'total': total
    }

def ganti_integer(perintah):
    while True:
        try:
            int_value = int(input(perintah))  # Menerima input dari pengguna
            if int_value > 0:  # Memastikan nilai lebih besar dari 0
                return int_value  # Mengembalikan nilai jika valid
            else:
                print(warna("\nSilakan Masukkan Angka Yang Valid ( Lebih Besar Dari 0 ) !!!\n",'merah'))
        except ValueError:
            print(warna("\nSilakan Masukkan Angka Yang Valid!!!",'merah'))

def clear_screen():
    os.system('cls')

def waktu():
    indonesia_time = datetime.now(pytz.timezone('Asia/Jakarta'))
    return warna(indonesia_time.strftime('%Y-%m-%d %H:%M:%S %Z'), 'cyan')

def user(): 
    clear_screen()
    while True:
        print(warna("\n============================== | NAMA PEMBELI | ==============================\n", 'putih'))  # Mengubah warna judul
        nama_user = input(warna("Masukan Nama Pembeli : ", 'putih')).upper()
        
        if not nama_user.strip():
            print("\nInput Tidak Boleh Kosong!!! Silahkan Isi")
            clear_screen()
            continue

        # Input nomor telepon
        while True:
            no_telepon = input(warna("Masukkan No Telepon  : ", 'putih')).strip()
            if no_telepon.isdigit() and len(no_telepon) >= 10:  # Validasi nomor telepon
                break
            else:
                print(warna("\nNo Telepon Tidak Valid! Harus Angka dan Minimal 10 Digit.", 'merah'))

        # Input tanggal dan jam
        while True:
            try:
                tanggal_input = input(warna("Masukkan Tanggal (DD-MM-YYYY) : ", 'putih'))
                jam_input = input(warna("Masukkan Jam (HH:MM) : ", 'putih'))

                # Menggabungkan tanggal dan jam menjadi satu string
                datetime_input = f"{tanggal_input} {jam_input}"

                # Mem-parsing input menjadi objek datetime
                waktu = datetime.strptime(datetime_input, "%d-%m-%Y %H:%M")
                break  # Keluar dari loop jika parsing berhasil

            except ValueError:
                print(warna("\nFormat tanggal atau jam tidak valid! Silakan coba lagi.", 'merah'))
        
        return nama_user, no_telepon, waktu # Mengembalikan nama_user, no telepone dan waktu

# Misalkan akun adalah dictionary yang menyimpan data member
akun = {}
# Menyimpan nomor telepon yang sudah terdaftar
no_telepon_terdaftar = set()

def add_member():
    clear_screen()
    print(warna("\n============================== | REGISTRASI MEMBER | ==============================\n", 'putih'))  

    while True:
        nama = input(warna("Masukkan Nama Member : ", 'putih')).strip()

        if nama in akun:
            print(warna("\nNama Member Sudah Terdaftar!", 'merah'))
            continue
        
        if not nama:
            print(warna("\nNama Tidak Boleh Kosong!", 'merah')) 
            continue
            
        # Memasukkan nomor telepon
        while True:
            no_telepon = input(warna("Masukkan No Telepon  : ", 'putih')).strip()
            if not no_telepon:
                print(warna("\nNo Telepon Tidak Boleh Kosong!", 'merah'))
                continue
            
            # Cek apakah nomor telepon sudah terdaftar
            if no_telepon in no_telepon_terdaftar:
                print(warna("\nNo Telepon Sudah Digunakan!", 'merah'))
                continue
            
            if not no_telepon.isdigit() or len(no_telepon) < 10:
                print(warna("\nNo Telepon Tidak Valid! Harus Angka dan Minimal 10 Digit.", 'merah'))
                continue
            break
        
        # Jika nomor telepon valid dan belum terdaftar, tambahkan ke set
        no_telepon_terdaftar.add(no_telepon)
        userid = generate_userid()
        discount_rate = random.randint(10, 20)
        
        akun[nama] = {
            'userid': userid,
            'discount_rate': discount_rate,
            'no_telepon': no_telepon
        }

        print(warna("\nMember Berhasil Didaftarkan!", 'hijau'))
        print(warna("==================================================================================", 'putih'))  
        print(warna(f"NAMA MEMBER         : {nama}", 'cyan'))
        print(warna(f"User ID Anda        : {userid}", 'cyan'))
        print(warna(f"No Telepon          : {no_telepon}", 'cyan'))
        print(warna("==================================================================================", 'putih'))  
        print(warna(f"Selamat!!! Anda Mendapatkan Diskon Sebesar {discount_rate}%", 'hijau'))  
        print(warna("==================================================================================", 'putih'))  

        # Simpan data ke Excel
        save_member(nama, userid, no_telepon, discount_rate)
        return nama

# Tentukan nama folder dan file
folder_path = 'DATA MEMBER'
file_excel = os.path.join(folder_path, 'data_member.xlsx') 

# Buat folder jika belum ada
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

def save_member(nama, userid, no_telepon, discount_rate):
    # Cek apakah file Excel sudah ada
    if not os.path.exists(file_excel):
        # Jika tidak ada, buat workbook baru dan tambahkan header
        wb = Workbook()
        ws = wb.active
        ws.append(['Nama', 'User ID', 'No Telepon', 'Discount Rate'])
    else:
        # Jika ada, buka workbook yang sudah ada
        wb = load_workbook(file_excel)
        ws = wb.active

    # Tambahkan data member baru
    ws.append([nama, userid, no_telepon, discount_rate])
    
    # Simpan workbook
    wb.save(file_excel)

def login_member():
    load_members()
    clear_screen()
    print(warna("\n============================= | CEK MEMBER | =============================\n", 'putih')) 

    max_attempts = 3  # Maksimal percobaan login
    attempts = 0      # Menghitung jumlah percobaan

    while attempts < max_attempts:

        nama = input(warna("Masukkan Nama Member : ", 'putih')).strip()
        no_telepon = input(warna("Masukkan No Telepon  : ", 'putih')).strip()

        if nama in akun and akun[nama] and akun[nama]['no_telepon'] == no_telepon:
            print(warna("\nLogin Berhasil!", 'hijau')) 
            print(warna("========================================================================", 'putih'))  
            print(warna(f"NAMA MEMBER          : {nama} ",'cyan'))  
            print(warna(f"NO TELEPON           : {no_telepon} ",'cyan'))  
            print(warna(f"UserID               : {akun[nama]['userid']} ",'cyan'))  
            print(warna("========================================================================", 'putih'))
            print(warna(f"Anda Mendapatkan Diskon Sebesar {akun[nama]['discount_rate']}%", 'cyan'))
            print(warna("========================================================================", 'putih'))
            break
        else:
            attempts += 1  # Tambah jumlah percobaan jika login gagal
            print(warna("\nNama, UserID, atau No Telepon Salah!", 'merah'))
            print(warna(f"Kesempatan tersisa: {max_attempts - attempts}", 'kuning'))

    if attempts == max_attempts:
        print(warna("\nAnda telah mencapai batas maksimum percobaan login. Silakan coba lagi nanti.", 'merah'))

    while True:
        kembali = input(warna('\ningin kembali ke menu utama? [ Y ]: ', 'biru')).lower()
        if kembali == 'y':
            print(warna(f"\nTERIMA KASIH TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru'))
            break 
        else:
            print(warna("Invalid Silahkan Masukan [ Y ]", 'merah'))
                        
def cek_member():
    load_members()
    clear_screen()
    print(warna("\n============================= | CEK MEMBER | =============================\n", 'putih')) 

    max_attempts = 3  # Maksimal percobaan login
    attempts = 0      # Menghitung jumlah percobaan

    while attempts < max_attempts:

        nama = input(warna("Masukkan Nama Member : ", 'putih')).strip()
        no_telepon = input(warna("Masukkan No Telepon  : ", 'putih')).strip()
        userid = input(warna("Masukkan UserID      : ", 'putih')).strip().upper()

        if nama in akun and akun[nama]['userid'] == userid and akun[nama]['no_telepon'] == no_telepon:
            print(warna("\nLogin Berhasil!", 'hijau')) 
            print(warna("========================================================================", 'putih'))  
            print(warna(f"NAMA MEMBER          : {nama} ",'cyan'))  
            print(warna(f"NO TELEPON           : {no_telepon} ",'cyan'))  
            print(warna(f"Anda Mendapatkan Diskon Sebesar {akun[nama]['discount_rate']}%", 'cyan'))
            print(warna("========================================================================", 'putih'))
            return nama
        else:
            attempts += 1  # Tambah jumlah percobaan jika login gagal
            print(warna("\nNama, UserID, atau No Telepon Salah!", 'merah'))
            print(warna(f"Kesempatan tersisa: {max_attempts - attempts}", 'kuning'))

    if attempts == max_attempts:
        print(warna("\nAnda telah mencapai batas maksimum percobaan login. Silakan coba lagi nanti.", 'merah')) 

        while True:
            buat_member = input(warna("\nAPAKAH ANDA INGIN MEMBUAT MEMBER? [ Y / N ]: ", 'biru')).lower()  
            if buat_member == 'y':
                return add_member() 
            elif buat_member == 'n':
                clear_screen()
                print(warna("=========================== | PEMBAYARAN | ===========================",'putih'))
                return None
            else:
                print(warna("Input tidak valid. Silakan masukkan [ Y / N ]", 'merah'))
                
def menu():
    while True:
        clear_screen()
        print(warna("============================= | RESERVASI PAKET MENU | =============================", 'putih')) 
        print(warna("DAFTAR PAKET YANG TERSEDIA :", 'putih'))
        print(warna("1. ROMANCE NIGHT", 'kuning')) 
        print(warna("2. FAMILY PACKAGE", 'kuning')) 
        print(warna("3. VIP PACKAGE", 'kuning')) 
        print(warna("4. KELUAR", 'kuning')) 
        print(warna("=================================================================================", 'putih'))  

        choice = input(warna("PILIHAN ANDA : ", 'putih'))
        if choice == '4':
            print(warna('\nTERIMA KASIH TELAH MENGGUNAKAN LAYANAN KAMI!!!', 'biru'))
            break
         
        elif choice == '1':
            nama_user, no_telepon, waktu = user()
            clear_screen()

            print(warna(f'''             
=========================== | ROMANCE NIGHT PACKAGE | =============================
                    
SELAMAT DATANG {nama_user} 

NAMA PAKET YANG DI PILIH :
- ROMANCE NIGHT PACKAGE 

FASILITAS YANG DI DAPAT  :
- MENDAPATKAN 2 BANGKU
                            
- PAKET MAKANAN & JUICE & MINUMAN : 
===================================================================================                      
PAKET A                       PAKET B
===================================================================================
- MAKANAN                     - MAKANAN
2 ( CHICKEN COLDEBLU )        2 ( FISH & CHIP )    
- MINUMAN                     - MINUMAN         
2 ( RUM )                     2 ( BEER )
===================================================================================  
            HARGA : Rp.375.000            HARGA : Rp.365.000     
===================================================================================
                            ''', 'putih'))
            
            print(warna("PILIHAN : \n", 'hijau'))
            print(warna('[ A ] , [ B ]', 'hijau'))

            while True:    
                pilih_paket = input(warna("\nPILIH PAKET : [ A ] | [ B ]: ", 'hijau')).lower()

                if pilih_paket == 'a':
                    print(warna("\n==================== PAKET A ====================", 'putih'))
                    print(warna("- 2 BANGKU", 'hijau'))
                    print(warna("- 2 CHICKEN COLDEBLU", 'hijau'))
                    print(warna("- 2 RUM", 'hijau'))
                    harga = 375000
                    paket = "PAKET A"
                    kode_pembayaran = "RMC-A" 
                    print(warna("\n=================================================", 'putih'))
                    print(warna(f"HARGA PAKET A : Rp.{harga:,.0f}",'hijau'))
                    print(warna("\n=================================================", 'putih'))
                elif pilih_paket == 'b':
                    print(warna("\n==================== PAKET B ====================", 'putih'))
                    print(warna("- 2 BANGKU", 'hijau'))
                    print(warna("- 2 FISH & CHIP", 'hijau'))
                    print(warna("- 2 BEER", 'hijau'))
                    harga = 365000
                    paket = "PAKET B"
                    kode_pembayaran = "RMC-B" 
                    print(warna("\n=================================================", 'putih'))
                    print(warna(f"HARGA PAKET B : Rp.{harga:,.0f}",'hijau'))
                    print(warna("\n=================================================", 'putih'))
                else:
                    print(warna("INVALID SILAHKAN MASUKAN SESUAI PAKET MENU!!!", 'merah'))
                    continue

                while True:
                    try:
                        choice = input(warna("\nAPAKAH ANDA MEMPUNYAI MEMBER?? [ Y / N ] : ", 'biru')).lower()
                        pilihan = "ROMANCE NIGHT PACKAGE" 

                        if choice == 'y':
                            nama_member = cek_member()
                            if nama_member:
                                total_harga = hitung_total(harga, nama_member)
                                print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                break

                        elif choice == 'n':
                            total_harga = hitung_total(harga)
                            while True:
                                buat_member = input(warna("\nAPAKAH ANDA INGIN MEMBUAT MEMBER??[ Y / N ] : ", 'biru')).lower()
                                if buat_member == 'y':
                                    nama_member = add_member() 
                                    print(warna("Member Berhasil Dibuat!", 'hijau'))
                                    total_harga = hitung_total(harga, nama_member)
                                    print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                    break
                                elif buat_member == 'n':
                                    break
                                else:
                                    print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))
                            break

                        else:
                            print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))

                    except Exception as e:
                        print(warna(f"Terjadi kesalahan: {e}", 'merah'))

                while True:
                    print(warna(f'''
NOTE INPUT PEMBAYARAN TIDAK BOLEH MENGGUNAKAN TITIK !!!
TOTAL HARGA YANG HARUS DI BAYAR
SUDAH TERMASUK DISKON, PAJAK DAN SERVICE : Rp.{total_harga['total']:,.0f}
''', 'kuning'))    
                    
                    input_uang = input(warna("        Masukan Uang Pembayaran          : Rp.",'hijau')).strip()
                    if input_uang.isdigit():
                        uang = int(input_uang)
                        print(warna("\nProses Pembayaran Sedang Dilakukan.....", 'putih'))
                        if uang >= total_harga['total']:
                            uang_kembali = uang - total_harga['total']
                            print(warna("Pembayaran berhasil!\n", 'hijau'))

                            # Simpan nota ke Excel
                            simpan_nota_paket(nama_user,no_telepon, waktu, pilihan, kode_pembayaran, total_harga)

                            print("======================================================")
                            print(warna(f"TANGGAL & JAM    : {waktu}", 'cyan'))
                            print(warna(f"NAMA             : {nama_user}", 'cyan'))
                            print(warna(f"NO TELEPON       : {no_telepon}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PESANAN                  ", 'putih'))
                            print("======================================================")
                            print(warna(f"PAKET            : {pilihan}", 'cyan'))
                            print(warna(f"TIPE             : {paket}", 'cyan'))
                            print(warna(f"KODE             : {kode_pembayaran}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PEMBAYARAN                  ", 'putih'))
                            print("======================================================")
                            print(warna(f"SUBTOTAL         : Rp.{total_harga['subtotal']:,.0f} ", 'kuning'))
                            print(warna(f"DISKON           : Rp.{total_harga['diskon']:,.0f}", 'kuning'))
                            print(warna(f"PAJAK            : Rp.{total_harga['pajak']:,.0f} ", 'kuning'))
                            print(warna(f"BIAYA SERVICE    : Rp.{total_harga['service']:,.0f} ", 'kuning'))
                            print(warna(f"TOTAL PEMBAYARAN : Rp.{total_harga['total']:,.0f} ", 'hijau'))
                            print("======================================================") 
                            print(warna(f"UANG USER        : Rp.{uang:,.0f} ", 'hijau'))
                            print(warna(f"UANG KEMBALI     : Rp.{uang_kembali:,.0f}", 'merah'))
                            print("======================================================")
                        else:
                            print(warna("Pembayaran Gagal. Silakan Masukan Uang Sesuai Pembayaran", 'merah'))
                            continue
                    else:
                        print(warna("Masukkan Jumlah Uang Yang Sesuai.", 'merah'))
                        continue 
                
                    while True:
                        kembali = input(warna('\ningin kembali ke menu utama? [ Y / N ]: ', 'biru')).lower()
                        if kembali == 'y':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru'))
                            menu_utama()  
                        elif kembali == 'n':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu paket...', 'biru'))
                            return menu() 
                        else:
                            print(warna("Invalid Silahkan Masukan [ Y / N ]", 'merah'))
       
        elif choice == '2':
            nama_user, no_telepon, waktu = user()
            clear_screen()

            print(warna(f'''
===================================== | FAMILY PACKAGE | =======================================

SELAMAT DATANG {nama_user} 

NAMA PAKET YANG DI PILIH :
- FAMILY PACKAGE 

FASILITAS YANG DI DAPAT : 
- MENDAPATKAN 4 BANGKU 

- PAKET MAKANAN & JUICE & MINUMAN : 
================================================================================================                            
   PAKET A                       PAKET B                        PAKET C
================================================================================================                
- MAKANAN                     - MAKANAN                      - MAKANAN
  4 ( CROISANT )                4 ( CROISANT )                 4 ( CROISANT )
  2 ( FISH & CHIP )             2 ( FISH & CHIP )              2 ( FISH & CHIP )
  2 ( LASAGNA )                 2 ( CARBONARA )                2 ( CACIO E PEPE )
- JUICE                       - JUICE                        - JUICE
  4 ( ORANGE JUICE )            4 ( APPLE JUICE )              4 ( MANGO JUICE )
- MINUMAN                     - MINUMAN                      - MINUMAN
  3 ( RUM )                     3 ( BEER )                     3 ( WATER )
===============================================================================================
HARGA : Rp.800.000            HARGA : Rp.805.000             HARGA : Rp.700.000
===============================================================================================
''', 'putih')) 
           
            print(warna("PILIHAN : \n", 'hijau'))
            print(warna('[ A ] , [ B ] , [ C ]', 'hijau'))

            while True:
                # IMPUT MAKANAN
                pilih_paket = input(warna("\nPILIH PAKET : [ A ] [ B ] [ C ] : ", 'hijau')).lower()
                if pilih_paket == 'a':
                    print(warna("\n==================== PAKET A ====================", 'putih'))
                    print("- 4 BANGKU")
                    print("- 4 CROISANT ")
                    print("- 2 FISH & CHIP ")
                    print("- 2 LASAGA ")
                    print("- 4 ORANGE JUICE ")
                    print("- 3 RUM")
                    harga = 800000
                    paket = ("PAKET A")
                    kode_pembayaran = "FMLY-A" 
                    print("=================================================")
                    print(warna(f"HARGA PAKET A : Rp.{harga:,.0f}",'hijau'))
                    print("=================================================")
                elif pilih_paket == 'b':
                    print(warna("\n==================== PAKET B  ====================", 'putih'))
                    print("- 4 BANGKU")
                    print("- 4 CROISANT ")
                    print("- 2 FISH & CHIP ")
                    print("- 2 CARBONARA ")
                    print("- 4 APPLE JUICE ")
                    print("- 3 BEER ")
                    harga = 805000
                    paket = ("PAKET B")
                    kode_pembayaran = "FMLY-B" 
                    print("=================================================")
                    print(warna(f"HARGA PAKET B : Rp.{harga:,.0f}",'hijau'))
                    print("=================================================")
                elif pilih_paket == 'c':
                    print(warna("\n==================== PAKET C ====================", 'putih'))
                    print("- 4 BANGKU")
                    print("- 4 CROISANT ")
                    print("- 2 FISH & CHIP ")
                    print("- 2 CACIO E PEPE")
                    print("- 4 MANGO JUICE ")
                    print("- 3 WATER ")
                    harga = 700000
                    paket = ("PAKET C")
                    kode_pembayaran = "FMLY-C" 
                    print("=================================================")
                    print(warna(f"HARGA PAKET C : Rp.{harga:,.0f}",'hijau'))                 
                    print("=================================================")
                else:
                    print(warna("INVALID SILAHKAN MASUKAN SESUAI KODE MENU!!!", 'merah'))
                    continue

                while True:
                    try:
                        choice = input(warna("\nAPAKAH ANDA MEMPUNYAI MEMBER?? [ Y / N ] : ", 'biru')).lower()
                        pilihan = "FAMILY PACKAGE" 

                        if choice == 'y':
                            nama_member = cek_member()
                            if nama_member:
                                total_harga = hitung_total(harga, nama_member)
                                print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                break

                        elif choice == 'n':
                            total_harga = hitung_total(harga)
                            while True:
                                try:
                                    buat_member = input(warna("\nAPAKAH ANDA INGIN MEMBUAT MEMBER??[ Y / N ] : ", 'biru')).lower()
                                    if buat_member == 'y':
                                        nama_member = add_member() 
                                        print(warna("Member Berhasil Dibuat!", 'hijau'))
                                        total_harga = hitung_total(harga, nama_member)
                                        print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                        break
                                    elif buat_member == 'n':
                                        break
                                    else:
                                        print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))
                                except Exception as regist:
                                    print(warna(f"Terjadi kesalahan: {regist}", 'merah'))
                            break
                        else:
                            print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))
                    except Exception as member:
                        print(warna(f"Terjadi kesalahan: {member}", 'merah'))

                while True:
                    print(warna(f'''
NOTE INPUT PEMBAYARAN TIDAK BOLEH MENGGUNAKAN TITIK !!!
TOTAL HARGA YANG HARUS DI BAYAR
SUDAH TERMASUK DISKON, PAJAK DAN SERVICE : Rp.{total_harga['total']:,.0f}
''', 'kuning'))    
                    input_uang = input(warna("        Masukan Uang Pembayaran          : Rp.",'hijau')).strip()
                    if input_uang.isdigit():
                        uang = int(input_uang)
                        print(warna("\nProses Pembayaran Sedang Dilakukan.....", 'putih'))
                        if uang >= total_harga['total']:
                            uang_kembali = uang - total_harga['total']
                            print(warna("Pembayaran berhasil!\n", 'hijau'))

                            # Simpan nota ke Excel
                            simpan_nota_paket(nama_user, no_telepon,waktu, pilihan, kode_pembayaran, total_harga)

                            print("======================================================")
                            print(warna(f"TANGGAL & JAM    : {waktu}", 'cyan'))
                            print(warna(f"NAMA             : {nama_user}", 'cyan'))
                            print(warna(f"NO TELEPON       : {no_telepon}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PESANAN                  ", 'putih'))
                            print("======================================================")
                            print(warna(f"PAKET            : {pilihan}", 'cyan'))
                            print(warna(f"TIPE             : {paket}", 'cyan'))
                            print(warna(f"KODE             : {kode_pembayaran}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PEMBAYARAN                  ", 'putih'))
                            print("======================================================")
                            print(warna(f"SUBTOTAL         : Rp.{total_harga['subtotal']:,.0f} ", 'kuning'))
                            print(warna(f"DISKON           : Rp.{total_harga['diskon']:,.0f}", 'kuning'))
                            print(warna(f"PAJAK            : Rp.{total_harga['pajak']:,.0f} ", 'kuning'))
                            print(warna(f"BIAYA SERVICE    : Rp.{total_harga['service']:,.0f} ", 'kuning'))
                            print(warna(f"TOTAL PEMBAYARAN : Rp.{total_harga['total']:,.0f} ", 'hijau'))
                            print("======================================================") 
                            print(warna(f"UANG USER        : Rp.{uang:,.0f} ", 'hijau'))
                            print(warna(f"UANG KEMBALI     : Rp.{uang_kembali:,.0f}", 'merah'))
                            print("======================================================")
                        else:
                            print(warna("Pembayaran Gagal. Silakan Masukan Uang Sesuai Pembayaran", 'merah'))
                            continue
                    else:
                        print(warna("Masukkan Jumlah Uang Yang Sesuai.", 'merah'))
                        continue 
                
                    while True:
                        kembali = input(warna('\ningin kembali ke menu utama? [ Y / N ]: ', 'biru')).lower()
                        if kembali == 'y':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru'))
                            menu_utama()  
                        elif kembali == 'n':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu paket...', 'biru'))
                            return menu() 
                        else:
                            print(warna("Invalid Silahkan Masukan [ Y / N ]", 'merah'))

        elif choice == '3':
            nama_user, no_telepon, waktu = user()
            clear_screen()
            print(warna(f'''
=========================== | VIP PACKAGE | =============================

SELAMAT DATANG {nama_user} 

NAMA PAKET YANG DI PILIH :
- VIP PACKAGE 

FASILITAS YANG DI DAPAT :
- MENDAPATKAN 1 BANGKU

- PAKET MAKANAN & JUICE & MINUMAN : 
=========================================================================                      
   PAKET A                       PAKET B
=========================================================================
- MAKANAN                     - MAKANAN
  1 ( FRENCH ONION SOUP )       1 ( FRENCH ONION SOUP )
  1 ( RAVIOLI )                 1 ( LASAGNA )
- MINUMAN                     - MINUMAN         
  1 ( MOJITO BLUE OCEAN )       1 ( VANILLA LATE )
=========================================================================  
HARGA : Rp.250.000            HARGA : Rp.275.00    
=========================================================================
                  ''', 'putih'))
            
            print(warna("PILIHAN : \n", 'hijau'))
            print(warna('[ A ] , [ B ]', 'hijau'))

            while True:    
                pilih_paket = input(warna("\nPILIH PAKET : [ A ] | [ B ]: ", 'hijau')).lower()

                if pilih_paket == 'a':
                    print(warna("\n==================== PAKET A ====================", 'putih'))
                    print(warna("- 1 BANGKU",'hijau'))
                    print(warna("- 1 FRENCH ONION SOUP", 'hijau'))
                    print(warna("- 1 RAVIOLI", 'hijau'))
                    print(warna("- 1 MOJITO BLUE OCEAN", 'hijau'))
                    harga = 250000
                    paket = ("PAKET A")
                    kode_pembayaran = "VIP-A" 
                    print("=================================================")
                    print(warna(f"HARGA PAKET A : Rp.{harga:,.0f}",'hijau'))
                    print("=================================================")
                elif pilih_paket == 'b':
                    print(warna("\n==================== PAKET B ====================", 'putih'))
                    print(warna("- 1 BANGKU", 'hijau'))
                    print(warna("- 1 FRENCH ONION SOUP", 'hijau'))
                    print(warna("- 1 LASAGNA", 'hijau'))
                    print(warna("- 1 VANILLA LATTE", 'hijau'))
                    harga = 275000
                    paket = ("PAKET B")
                    kode_pembayaran = "VIP-B" 
                    print("=================================================")
                    print(warna(f"HARGA PAKET B : Rp.{harga:,.0f}",'hijau'))
                    print("=================================================")
                else :
                    print(warna("INVALID SILAHKAN MASUKAN SESUAI PAKET MENU!!!", 'merah'))
                    continue

                while True:
                    try:
                        choice = input(warna("\nAPAKAH ANDA MEMPUNYAI MEMBER?? [ Y / N ] : ", 'biru')).lower()
                        pilihan = "VIP PACKAGE" 

                        if choice == 'y':
                            nama_member = cek_member()
                            if nama_member:
                                total_harga = hitung_total(harga, nama_member)
                                print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                break

                        elif choice == 'n':
                            total_harga = hitung_total(harga)
                            while True:
                                buat_member = input(warna("\nAPAKAH ANDA INGIN MEMBUAT MEMBER??[ Y / N ] : ", 'biru')).lower()
                                if buat_member == 'y':
                                    nama_member = add_member() 
                                    print(warna("Member Berhasil Dibuat!", 'hijau'))
                                    total_harga = hitung_total(harga, nama_member)
                                    print(warna(f"\nSELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{total_harga['diskon']:,.0f} !!!", 'hijau')) 
                                    break
                                elif buat_member == 'n':
                                    break
                                else:
                                    print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))
                            break

                        else:
                            print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))

                    except Exception as e:
                        print(warna(f"Terjadi kesalahan: {e}", 'merah'))

                while True:
                    print(warna(f'''
NOTE INPUT PEMBAYARAN TIDAK BOLEH MENGGUNAKAN TITIK !!!
TOTAL HARGA YANG HARUS DI BAYAR
SUDAH TERMASUK DISKON, PAJAK DAN SERVICE : Rp.{total_harga['total']:,.0f}
''', 'kuning'))    
                    input_uang = input(warna("        Masukan Uang Pembayaran          : Rp.",'hijau')).strip()
                    if input_uang.isdigit():
                        uang = int(input_uang)
                        print(warna("\nProses Pembayaran Sedang Dilakukan.....", 'putih'))
                        if uang >= total_harga['total']:
                            uang_kembali = uang - total_harga['total']
                            print(warna("Pembayaran berhasil!\n", 'hijau'))

                            # Simpan nota ke Excel
                            simpan_nota_paket(nama_user,no_telepon, waktu, pilihan, kode_pembayaran, total_harga)

                            print("======================================================")
                            print(warna(f"TANGGAL & JAM    : {waktu}", 'cyan'))
                            print(warna(f"NAMA             : {nama_user}", 'cyan'))
                            print(warna(f"NO TELEPON       : {no_telepon}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PESANAN                  ", 'putih'))
                            print("======================================================")
                            print(warna(f"PAKET            : {pilihan}", 'cyan'))
                            print(warna(f"TIPE             : {paket}", 'cyan'))
                            print(warna(f"KODE             : {kode_pembayaran}", 'cyan'))
                            print("======================================================")
                            print(warna("                  RINCIAN PEMBAYARAN                ", 'putih'))
                            print("======================================================")
                            print(warna(f"SUBTOTAL         : Rp.{total_harga['subtotal']:,.0f} ", 'kuning'))
                            print(warna(f"DISKON           : Rp.{total_harga['diskon']:,.0f}", 'kuning'))
                            print(warna(f"PAJAK            : Rp.{total_harga['pajak']:,.0f} ", 'kuning'))
                            print(warna(f"BIAYA SERVICE    : Rp.{total_harga['service']:,.0f} ", 'kuning'))
                            print(warna(f"TOTAL PEMBAYARAN : Rp.{total_harga['total']:,.0f} ", 'hijau'))
                            print("======================================================") 
                            print(warna(f"UANG USER        : Rp.{uang:,.0f} ", 'hijau'))
                            print(warna(f"UANG KEMBALI     : Rp.{uang_kembali:,.0f}", 'merah'))
                            print("======================================================")
                        else:
                            print(warna("Pembayaran Gagal. Silakan Masukan Uang Sesuai Pembayaran", 'merah'))
                            continue
                    else:
                        print(warna("Masukkan Jumlah Uang Yang Sesuai.", 'merah'))
                        continue 
                
                    while True:
                        kembali = input(warna('\ningin kembali ke menu utama? [ Y / N ]: ', 'biru')).lower()
                        if kembali == 'y':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru'))
                            menu_utama()  
                        elif kembali == 'n':
                            print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru'))
                            input(warna('Tekan Enter untuk kembali ke menu paket...', 'biru'))
                            return menu() 
                        else:
                            print(warna("Invalid Silahkan Masukan [ Y / N ]", 'merah'))
                        
# Fungsi untuk menyimpan nota ke dalam file Excel
def simpan_nota_paket(nama_user, no_telepon, waktu, paket, kode_pembayaran, total_harga):
    folder_path = 'DATA PEMESANAN RESERVASI PAKET'
    file_excel = os.path.join(folder_path, 'pesanan reservasi paket.xlsx')

    # Buat folder jika belum ada
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    if not os.path.exists(file_excel):
    # Jika file belum ada, buat file baru dan tambahkan header
        wb = Workbook()
        ws = wb.active
        # Menambahkan header
        ws.append(['Nama', 'No Telepon' ,'Tanggal & Waktu','Paket', 'Kode Pembayaran', 'Subtotal', 'Diskon', 'Pajak', 'Biaya Service', 'Total Pembayaran' ])
    else:
        # Jika file sudah ada, buka file yang ada
        wb = load_workbook(file_excel)
        ws = wb.active

    # Menambahkan data nota
    data = [
        nama_user,
        no_telepon,
        waktu,
        paket,
        kode_pembayaran,
        total_harga['subtotal'],
        total_harga['diskon'],
        total_harga['pajak'],
        total_harga['service'],
        total_harga['total']
    ]
    ws.append(data)

    # Simpan workbook
    wb.save(file_excel)

def pilih(nama_user):
    while True:
            time = waktu()
            clear_screen()
            print("==========================================================================\n")
            print(f"SELAMAT DATANG {warna(nama_user, 'biru')}  ")
            print(time)
            print('''
===================== | PILIH MAKANAN & MINUMAN | ========================
- MAKANAN & JUICE & MINUMAN : 
==========================================================================                     
  MAKANAN              HARGA         MINUMAN                HARGA         
==========================================================================

1.CHICKEN COLDEBLUE   Rp.50.000      A.MOJITO BLUE OCEAN   Rp.15.000     
2.FRENCH ONION SOUP   Rp.50.000      B.MILKSHAKE           Rp.20.000       
3.FISH & CHIP         Rp.40.000      C.VANILLA LATE        Rp.25.000       
4.FRENCH FRIES        Rp.15.000      D.WATER               Rp.20.000      
5.LASAGNA             Rp.70.000      E.RUM                 Rp.150.000                            
6.CABONARA            Rp.75.000      F.BEER                Rp.150.000                            
                                                                                             
===========================================================================
                  ''')

            jenis_pesanan = []
            pesanan = ganti_integer(" PESANAN : ")
            print(warna("\nSILAHKAN MEMASUKAN KODE PESANAN!!!\n",'hijau'))
            for i in range(pesanan):
                pesanan_ke = i + 1
                print(f"PESANAN KE - {pesanan_ke}")
                    
                while True:
                    makanan = input(warna("PILIH MENU MAKANAN [1/2/3/4/5/6] : ",'hijau')).upper()
                    if makanan not in ['1', '2', '3', '4', '5', '6']: 
                        print(warna("Kode tidak valid. Silakan masukkan dengan benar",'merah'))
                        continue
                    
                    banyak_pesanan_makanan = ganti_integer(warna("BANYAK PESANAN MAKANAN : ",'hijau'))

                    while True:
                        minuman = input(warna("\nPILIH MENU MINUMAN [A/B/C/D/E/F] : ",'hijau')).upper()
                        if minuman not in ['A', 'B', 'C', 'D', 'E', 'F']:
                            print(warna("Kode tidak valid. Silakan masukkan dengan benar",'merah'))
                            continue
        
                        banyak_pesanan_minuman = ganti_integer(warna("BANYAK PESANAN MINUMAN : ", 'hijau'))
 
                        jenis_pesanan.append((makanan, minuman, banyak_pesanan_makanan, banyak_pesanan_minuman))
                        break  # Exit juice selection loop
                    break  # Exit minuman selection loop
            return jenis_pesanan

def total(jenis_pesanan):
    total_harga = 0
    rincian_pesanan = []

    for makanan, minuman, banyak_pesanan_makanan, banyak_pesanan_minuman in jenis_pesanan:        
        # Harga dan nama makanan
        if makanan in ['1','2']:
           harga_makanan = 50000
           nama_makanan = 'CHICKEN COLDEBLUE' if makanan == '1' else 'FRENCH ONION SOUP'
        elif makanan == '3':
           harga_makanan = 40000
           nama_makanan = 'FISH & CHIP'
        elif makanan == '4':
           harga_makanan = 15000
           nama_makanan = 'FRENCH FRIES'
        elif makanan == '5':
           harga_makanan = 70000
           nama_makanan = 'LASAGNA'
        elif makanan == '6':
           harga_makanan = 75000
           nama_makanan = 'CABONARA'
        else:
            harga_makanan = 0
            nama_makanan = ''

        # Harga dan nama minuman
        if minuman in ['A', 'B', 'C', 'D', 'E', 'F']:
           harga_minuman = {'A': 15000, 'B': 20000, 'C': 25000, 'D': 20000, 'E': 150000, 'F': 150000}[minuman]
           nama_minuman = {'A': 'MOJITO BLUE OCEAN', 'B': 'MILKSHAKE', 'C': 'VANILLA LATE', 'D': 'WATER', 'E': 'RUM', 'F': 'BEER'}[minuman]
        else:
            harga_minuman = 0
            nama_minuman = "" 

        # Tambahkan ke rincian pesanan
        rincian_pesanan.append((nama_makanan, nama_minuman, banyak_pesanan_makanan, banyak_pesanan_minuman))
        
        # Hitung total harga
        total_harga += (harga_makanan * banyak_pesanan_makanan) + \
                       (harga_minuman * banyak_pesanan_minuman)

    return total_harga, harga_makanan, harga_minuman, rincian_pesanan
    
def pembayaran(jenis_pesanan, nama_user, no_telepon, waktu):
    total_harga, harga_makanan, harga_minuman, rincian_pesanan = total(jenis_pesanan)
    diskon = 0
    
    while True:
        choice = input(warna("\nAPAKAH ANDA MEMPUNYAI MEMBER?? [ Y / N ] : ", 'biru')).lower()
        pajak = total_harga * 0.05
        service = total_harga * 0.02
 
        if choice == 'y':
            nama_member = cek_member()
            if nama_member:
                diskon = total_harga * (akun[nama_member]['discount_rate'] / 100)
                total_harga_pilih = total_harga + pajak + service - diskon 
                print(warna(f"SELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{diskon:,.0f} !!!", 'hijau')) 
                break
            
        elif choice == 'n':
            while True:
                registrasi_member = input(warna("\nAPAKAH ANDA INGIN MEMBUAT MEMBER?? [ Y / N ] : ",'biru')).lower()
                if registrasi_member == 'y':
                    nama_member = add_member()
                    diskon = total_harga * (akun[nama_member]['discount_rate'] / 100)
                    total_harga_pilih = total_harga + pajak + service - diskon 
                    print(warna(f"SELAMAT ANDA MENDAPATKAN DISKON SEBESAR Rp.{diskon:,.0f} !!!", 'hijau')) 
                    break
                elif registrasi_member == 'n':
                    total_harga_pilih = total_harga + pajak + service
                    break
                else:
                    print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))
            break
        else:
            print(warna("INVALID SILAHKAN MASUKAN  [ Y / N ]", 'merah'))

    total_harga_pilih = total_harga + pajak + service - diskon

    while True:
        print(warna(f'''
NOTE INPUT PEMBAYARAN TIDAK BOLEH MENGGUNAKAN TITIK !!!   
TOTAL HARGA YANG HARUS DI BAYAR
SUDAH TERMASUK DISKON, PAJAK DAN SERVICE : Rp.{total_harga_pilih:,.0f}
''','kuning'))  
        input_uang = input(warna("        Masukan Uang Pembayaran          : Rp.",'hijau')).strip()
        if input_uang.isdigit():
            uang = int(input_uang)
            print("\nProses Pembayaran Sedang Dilakukan.....")
            if uang >= total_harga_pilih:
                uang_kembali = uang - total_harga_pilih 
                print(warna("Pembayaran berhasil!\n",'hijau'))

                simpan_nota_pilihan(nama_user, no_telepon, waktu, rincian_pesanan, harga_makanan, harga_minuman, diskon, pajak, service, total_harga_pilih)

                print("======================================================")
                print(warna(f"TANGGAL & JAM    : {waktu}",'cyan'))
                print(warna(f"NAMA             : {nama_user}",'cyan'))
                print(warna(f"NO TELEPON       : {no_telepon}",'cyan'))
                print("======================================================")
                print(warna("                  RINCIAN PESANAN                  ",'putih'))
                data_makanan = {
                    "": [], # Makanan
                    " ": [], # Banyak Pesanan
                    "  ": [], # Harga Satuan
                    "   ": [] # TOTAL
                }
                data_minuman = {
                    "": [], # Minuman
                    " ": [], # Banyak Pesanan
                    "  ": [], # Harga Satuan
                    "   ": [] # TOTAL
                }

                # HARGA MAKANAN DIKALI BANYAK PESANAN
                for nama_makanan, nama_minuman, banyak_pesanan_makanan, banyak_pesanan_minuman in rincian_pesanan:
                    harga_total_makanan = harga_makanan * banyak_pesanan_makanan
                    harga_total_minuman = harga_minuman * banyak_pesanan_minuman

                    data_makanan[""].append(nama_makanan)
                    data_makanan[" "].append(f"{banyak_pesanan_makanan}x")
                    data_makanan["  "].append(harga_makanan)
                    data_makanan["   "].append(f"Rp.{harga_total_makanan:,.0f}")


                    data_minuman[""].append(nama_minuman)
                    data_minuman[" "].append(f"{banyak_pesanan_minuman}x")
                    data_minuman["  "].append(harga_minuman)
                    data_minuman["   "].append(f"Rp.{harga_total_minuman:,.0f}")

                data_makanan_total = pd.DataFrame(data_makanan, index=range(1, len(rincian_pesanan) + 1))
                data_minuman_total = pd.DataFrame(data_minuman, index=range(1, len(rincian_pesanan) + 1))
             
                print("======================================================")
                print(warna(f"{data_makanan_total}",'hijau'))
                print(warna(f"{data_minuman_total}",'hijau'))
                print("\n======================================================")
                print(warna("                  RINCIAN PEMBAYARAN                  ",'putih'))
                print("======================================================")
                print(warna(f"SUB TOTAL        : Rp.{total_harga:,.0f}", 'kuning'))
                print(warna(f"DISKON           : Rp.{diskon:,.0f}", 'kuning'))
                print(warna(f"PAJAK            : Rp.{pajak:,.0f}", 'kuning'))
                print(warna(f"BIAYA SERVICE    : Rp.{service:,.0f}", 'kuning'))
                print(warna(f"TOTAL PEMBAYARAN : Rp.{total_harga_pilih:,.0f}", 'hijau'))
                print("======================================================") 
                print(warna(f"UANG USER        : Rp.{uang:,.0f} ",'hijau'))
                print(warna(f"UANG KEMBALI     : Rp.{uang_kembali:,.0f}",'merah'))
                print("======================================================") 
            else:
                print(warna("Pembayaran Gagal. Silakan Masukan Uang Sesuai Pembayaran",'merah'))
                continue
        else:
            print(warna("Masukkan Jumlah Uang Yang Sesuai.",'merah'))
            continue
        
        while True:       
            kembali = input(warna('\nIngin kembali ke menu utama? [ Y ]: ','biru')).lower()
            if kembali == 'y':
                print(warna(f"\nTERIMA KASIH {nama_user} TELAH MENGGUNAKAN LAYANAN KAMI !!!", 'biru')) 
                return None
            else:
                print(warna("Invalid Silahkan Masukan [ Y ]",'merah'))

# Fungsi untuk menyimpan nota ke dalam file Excel
def simpan_nota_pilihan(nama_user, no_telepon, waktu, rincian_pesanan, harga_makanan, harga_minuman, diskon, pajak, service, total_harga_pilih):
    folder_path = 'DATA PEMESANAN MENU'
    file_excel = os.path.join(folder_path, 'pesanan menu.xlsx')

    # Buat folder jika belum ada
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    if not os.path.exists(file_excel):
        # Jika file belum ada, buat file baru dan tambahkan header
        wb = Workbook()
        ws = wb.active
        # Menambahkan header
        ws.append(['Nama', 'No Telepon', 'Tanggal & Waktu', 'Makanan', 'Qty', 'Minuman', 'Qty', 'Subtotal', 'Diskon', 'Pajak', 'Biaya Service', 'Total Pembayaran'])
    else:
        # Jika file sudah ada, buka file yang ada
        wb = load_workbook(file_excel)
        ws = wb.active

    # Menambahkan data nota
    for item in rincian_pesanan:
        nama_makanan, nama_minuman, banyak_pesanan_makanan, banyak_pesanan_minuman = item
        # Hitung total harga untuk item ini
        total_harga_item = (harga_makanan * banyak_pesanan_makanan) + (harga_minuman * banyak_pesanan_minuman)
        data = [
            nama_user,
            no_telepon,
            waktu,
            nama_makanan,
            banyak_pesanan_makanan,
            nama_minuman,
            banyak_pesanan_minuman,
            total_harga_item,
            diskon,
            pajak,
            service,
            total_harga_pilih,
        ]
        ws.append(data)

    # Simpan workbook
    wb.save(file_excel)

def menu_utama(): 
    while True:
        clear_screen()
        print(warna("============================= | SELAMAT DATANG DI APLIKASI KAMI | =============================", 'putih'))
        print(warna("PELAYANAN KAMI :", 'putih'))
        print(warna("1. RESERVASI PAKET", 'kuning'))
        print(warna("2. RESERVASI MENU", 'kuning'))
        print(warna("3. REGISTRASI MEMBER", 'kuning'))
        print(warna("4. CEK MEMBER", 'kuning'))
        print(warna("5. KELUAR", 'kuning'))
        print(warna("===============================================================================================", 'putih'))

        choice = input(warna("PILIHAN ANDA : ", 'putih'))
        if choice == '5':
            input(warna('\nTERIMAKASIH SUDAH MENGGUNAKAN APLIKASI KAMI!!!!', 'biru')) 
            break
             
        elif choice == '1':
            menu()
            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru')) 

        elif choice == '2':
            # load_members()
            nama_user, no_telepon, waktu = user()
            jenis_pesanan = pilih(nama_user)
            total(jenis_pesanan)
            pembayaran(jenis_pesanan, nama_user, no_telepon, waktu)
            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru')) 

        elif choice == '3':
            add_member()
            input(warna('Tekan Enter untuk kembali ke menu utama...', 'biru')) 
  
        elif choice == '4':
            login_member()

        else:
            print("Pilihan tidak valid. Silakan coba lagi.")

menu_utama()